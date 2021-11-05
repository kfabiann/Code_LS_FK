"""

Modified: 04.10.2021
@author: Fabian Kneubuehler

Theme: Comparison of process forces 

#-*- coding:utf-8 -*-
"""

import numpy as np
import matplotlib.pyplot as plt
import matplotlib.cm as cm
import os
import sys
import openpyxl
from openpyxl import load_workbook
import numpy_indexed as npi
from matplotlib.lines import lineStyles


##########################################################################
##########################################################################
# General settings

# https://matplotlib.org/users/usetex.html
plt.rc('text', usetex=True)
plt.rc('font', family='serif')

# Float format constrained to two decimal places
float_formatter = "{:.2f}".format
np.set_printoptions(formatter={'float_kind': float_formatter})

##########################################################################
##########################################################################
# Load data

# Load Excel file with process parameters
filename_Excel = "Zerspanungsversuche_Hobeln_Protokoll_Gesamt_03102021.xlsx"
workbook = load_workbook(
    filename=filename_Excel, data_only=True)

MainSheet = workbook.worksheets[0]

# Path selection Windows:
Pfad = r"C:\Users\kfabian\eclipse-workspace\Kraftauswertung_v001\Evaluation_files"
# Pfad Mac:
# r"/Users/fabiankneubuhler/Desktop/Kraftrauswertung/Python/Daten/Gruppe1"

# Initalise list
Dateien = []
Experiments = []
Force = []

# Select Experiments to be evaluated

# Manual selection of Experiments for evaluation
#Experiments = [str("521_5"), str("521_6"), str("521_7"), str("522_1"), str("522_2"), str("522,3")]

# Selection of Experiments for evaluation via txt file
with open('ToEvaluate.txt') as ToEval:
    Experiments = ToEval.read().splitlines()

for r, d, f in os.walk(Pfad):
    for Datei in f:
        for i in range(len(Experiments)):
            if ('.txt' in Datei) and ((Experiments[i]) in Datei):
                Dateien.append(os.path.join(r, Datei))


##########################################################################
##########################################################################
# Functions

def GroupArray(arr, col):
    arr = npi.group_by(arr[:, col]).split(
        arr[:, [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]])
    return arr


def SortArray(arr):
    arr = arr[arr[:, 0].argsort()]
    return arr

##########################################################################
##########################################################################


for idx, Datei in enumerate(Dateien):

    Lines = np.genfromtxt(Datei, dtype='str')

    # Find row number of experiment
    for cell in MainSheet['A']:
        if cell.value == Lines[0]:
            ZeileExcel = cell.row

    # Find cutting speed
    vc = float(MainSheet.cell(row=ZeileExcel, column=15).value)
    # Find feed
    f = float(MainSheet.cell(row=ZeileExcel, column=12).value)
    # Find Wandstaerke
    Wandstaerke = float(MainSheet.cell(row=ZeileExcel, column=18).value) * 1000
    # Find # force measurements
    Repetitions = int(MainSheet.cell(row=ZeileExcel, column=16).value)

    # Initialise temporary array
    Force_temp = np.zeros((Repetitions, 10))

    # Split Experiment in order to store in array
    ExperimentNumber = Lines[0].split('_')

    for j in range(Repetitions):
        Force_temp[j, :] = [vc, f, Lines[1 + j], Lines[Repetitions + 1 + j], Lines[2 * Repetitions + 1 + j], Lines[3 * Repetitions +
                                                                                                                   1 + j], Lines[4 * Repetitions + 1 + j], Lines[5 * Repetitions + 1 + j], ExperimentNumber[0], ExperimentNumber[1]]

    if idx == 0:
        Force = np.array(Force_temp)

    else:
        Force = np.concatenate((Force, Force_temp), axis=0)


##########################################################################
##########################################################################
# Data evaluation
Length1 = len(Force)

# Delete rows that contain Cutting Force of 0
Force = Force[(Force[:, 2] != 0), :]

Length2 = len(Force)

print(str(Length1 - Length2) +
      ' cutting force measurements contained 0 N and were deleted!')

Force = SortArray(Force)  # Sort array by cutting speed

# Group array for the same cutting speeds
Force = GroupArray(Force, 0)

# Group arrays with the same cutting speeds into arrays with same feed
ArrLength = []  # List with number of repetitions, which does not need to be equal to the number in the excel sheet!

for i in range(len(Force)):
    # group array of the same cutting speed in groups of the same feed
    Force[i] = GroupArray(Force[i], 1)
    ArrLength.append(len(Force[i]))

ArrLength = sum(ArrLength)


# Plot grouped data and calculate mean force values and mean standard
# deviations
Force_eval = np.zeros((ArrLength, 19))

ii = 0  # Row index for evaluated values

for i in range(len(Force)):

    for j in range(len(Force[i])):
        # Calculations

        # Cutting force
        Force_eval[ii, 0] = Force[i][j][0, 0]  # cutting speed v_c
        Force_eval[ii, 1] = Force[i][j][0, 1]  # feed f
        Force_eval[ii, 2] = np.round(
            np.mean(Force[i][j][:, 2]), 1)  # Mean cutting force
        Force_eval[ii, 3] = np.round(
            np.std(Force[i][j][:, 2]), 2)  # Standard deviation
        Force_eval[ii, 4] = np.round(np.median(Force[i][j][:, 2]), 2)  # Median
        Force_eval[ii, 5] = np.round(np.quantile(
            Force[i][j][:, 2], 0.25), 2)  # Lower quantile
        Force_eval[ii, 6] = np.round(np.quantile(
            Force[i][j][:, 2], 0.75), 2)  # Upper quantile

        # Feed force
        Force_eval[ii, 7] = np.round(
            np.mean(Force[i][j][:, 4]), 1)  # Mean feed force
        Force_eval[ii, 8] = np.round(
            np.std(Force[i][j][:, 4]), 2)  # Standard deviation
        Force_eval[ii, 9] = np.round(np.median(Force[i][j][:, 4]), 2)  # Median
        Force_eval[ii, 10] = np.round(np.quantile(
            Force[i][j][:, 4], 0.25), 2)  # Lower quantile
        Force_eval[ii, 11] = np.round(np.quantile(
            Force[i][j][:, 4], 0.75), 2)  # Upper quantile

        # Passive force
        Force_eval[ii, 12] = np.round(
            np.mean(Force[i][j][:, 6]), 1)  # Mean feed force
        Force_eval[ii, 13] = np.round(
            np.std(Force[i][j][:, 6]), 2)  # Standard deviation
        Force_eval[ii, 14] = np.round(
            np.median(Force[i][j][:, 5]), 2)  # Median
        Force_eval[ii, 15] = np.round(np.quantile(
            Force[i][j][:, 6], 0.25), 2)  # Lower quantile
        Force_eval[ii, 16] = np.round(np.quantile(
            Force[i][j][:, 6], 0.75), 2)  # Upper quantile

        # Experiment no.
        Force_eval[ii, 17] = Force[i][j][0, 8]  # Experiment no.
        Force_eval[ii, 18] = Force[i][j][0, 9]  # Counter

        # Plot
        title = str(
            'Distribution of force measurements for $v_c$=' + str(Force[i][j][0, 0]) + ' and $f$=' + str(Force[i][j][0, 1]))
        fig, (ax1, ax2, ax3) = plt.subplots(1, 3, constrained_layout=True)
        # Assign different colors to each data point
        for jj in range(len(Force[i][j][:, 1])):
            colors = cm.rainbow(np.linspace(0, 1, len(Force[i][j][:, 1]) + 1))
            ax1.errorbar(Force[i][j][jj, 1], Force[i][j][jj, 2], Force[i][j][jj, 3], label=(str(int(
                Force[i][j][jj, 8])) + '\_' + str(int(Force[i][j][jj, 9]))), marker=".", markersize='7', color=colors[jj], linestyle='None', capsize=5)
            ax2.errorbar(Force[i][j][jj, 1], Force[i][j][jj, 4], Force[i][j][jj, 5], label=(str(int(Force[i][j][jj, 8])) +
                                                                                            '\_' + str(int(Force[i][j][jj, 9]))), marker=".", markersize='7', color=colors[jj], linestyle='None', capsize=5)
            ax3.errorbar(Force[i][j][jj, 1], Force[i][j][jj, 6], Force[i][j][jj, 7], label=(str(int(Force[i][j][jj, 8])) +
                                                                                            '\_' + str(int(Force[i][j][jj, 9]))), marker=".", markersize='7', color=colors[jj], linestyle='None', capsize=5)

        # Plot calculated mean values + SD
        boxprops = dict(linestyle='-', linewidth=1.5)
        medianprops = dict(linestyle='-', linewidth=1.5, color='firebrick')
        meanprops = dict(linestyle='--', linewidth=1.5)
        flierprops = dict(markeredgewidth=1.5, markersize=8)

        ax1.boxplot([Force[i][j][:, 2]], showmeans=True, meanline=True,  boxprops=boxprops, medianprops=medianprops, capprops=boxprops, whiskerprops=boxprops, flierprops=flierprops, meanprops=meanprops,
                    positions=[Force[i][j][0, 1]])
        # Mean, median, upper and lower quantile without boxplot
        # ax1.errorbar(Force_eval[ii, 1], Force_eval[ii, 2], Force_eval[ii, 3], label=(
        #     'Mean cutting force'), marker="o", color=colors[jj + 1], linestyle='None', capsize=10)
        # ax1.plot(Force_eval[ii, 1], Force_eval[ii, 4],
        #          label=('Median cutting force'), marker="o")
        # ax1.plot(Force_eval[ii, 1], Force_eval[ii, 5],
        #          label=('Median lower'), marker="o")
        # ax1.plot(Force_eval[ii, 1], Force_eval[ii, 6],
        #          label=('Median upper'), marker="o")

        ax2.boxplot([Force[i][j][:, 4]], showmeans=True, meanline=True, boxprops=boxprops, medianprops=medianprops, capprops=boxprops, whiskerprops=boxprops, flierprops=boxprops, meanprops=meanprops,
                    positions=[Force[i][j][0, 1]])
        # Mean, median, upper and lower quantile without boxplot
        #ax2.errorbar(Force_eval[ii, 1], Force_eval[ii, 4], Force_eval[ii, 5], label=('Mean feed force'), marker="o", color=colors[jj + 1], linestyle='None', capsize=10)
        # ax2.plot(Force_eval[ii, 1], Force_eval[ii, 9],
        #          label=('Median cutting force'), marker="o")
        # ax2.plot(Force_eval[ii, 1], Force_eval[ii, 10],
        #          label=('Median lower'), marker="o")
        # ax2.plot(Force_eval[ii, 1], Force_eval[ii, 11],
        #          label=('Median upper'), marker="o")

        ax3.boxplot([Force[i][j][:, 6]], showmeans=True, meanline=True, boxprops=boxprops, medianprops=medianprops, capprops=boxprops, whiskerprops=boxprops, flierprops=boxprops, meanprops=meanprops,
                    positions=[Force[i][j][0, 1]])

        # Plot settings
        ax1.set_title('Cutting Force')
        ax2.set_title('Feed Force')
        ax3.set_title('Passive Force')
        ax1.set_xlabel(r'$f$ [mm]')
        ax2.set_xlabel(r'$f$ [mm]')
        ax3.set_xlabel(r'$f$ [mm]')
        ax1.set_ylabel(r'Cutting force [N/mm]')
        ax2.set_ylabel(r'Feed force [N/mm]')
        ax3.set_ylabel(r'Passive force [N/mm]')
        fig.suptitle(title)
        # ax1.legend(loc="lower left")
        # ax2.legend(loc="lower left")
        #ax3.legend(loc="lower left")
        ax3.legend(loc='center left', bbox_to_anchor=(1, 0.5))

        FileName = r"\RepetionsForceEvaluation_vc" + \
            str(Force[i][j][0, 0]) + '_f' + str(Force[i][j][0, 1])
        FileName = FileName.replace(".", "_")

        plt.savefig(r"Figures\EPS\RepetionForceEvaluations" +
                    FileName + ".eps", dpi=300)
        plt.savefig(r"Figures\PNG\RepetionForceEvaluations" +
                    FileName + ".png", dpi=300)
        plt.show(block=True)

        ii = ii + 1


Head = str(
    'v_c [m/min], f [mm], F_c [N/mm], SD F_c [N/mm], F_f [N/mm], SD F_f [N/mm], Experiment no.')
DataToSave = np.column_stack([Force_eval])
Datafile_path = r"Evaluation_files\Evaluation_Repetions\ProcessForceEvaluationRepetions"
np.savetxt((Datafile_path + ".txt"), DataToSave,
           fmt='%s', delimiter=' ', header=Head)
np.save(Datafile_path, Force_eval)

print("Text file saved!")
