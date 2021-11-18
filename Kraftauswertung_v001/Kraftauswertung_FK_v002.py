"""
Script based on: "Auswertung_Prozeszkraefte_Zerspanungsversuche.py" @Hagen Klippel

Modified: 25.08.2021
@author: Fabian Kneubühler

Theme: Evaluation of process forces 

#-*- coding:utf-8 -*-
"""

import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as mtick
from matplotlib.widgets import RectangleSelector
import os
import openpyxl
from openpyxl import load_workbook

# https://matplotlib.org/users/usetex.html
plt.rc('text', usetex=True)
plt.rc('font', family='serif')

# Drift compensation
DriftCompensationExe = 1
DriftCompensation = 1
InitialFigure = 1


# Load Excel file
filename_Excel = "Zerspanungsversuche_Hobeln_Protokoll_Gesamt_03102021.xlsx"
workbook = load_workbook(
    filename=filename_Excel, data_only=True)

#sheet = workbook.active
#sheet = workbook.get_sheet_by_name('Overview')
MainSheet = workbook.worksheets[0]
EvalDetails = workbook.worksheets[1]

# Select data of Excel file
# firstRow = 8
# firstCol = 1
# nCols = 30
# nRows = 12
#
# allCells = np.array([[cell.value for cell in row]
#                      for row in MainSheet.iter_rows()])
#
#
# data = allCells[(firstRow - 1):(firstRow - 1 + nRows),
#                 (firstCol - 1):(firstCol - 1 + nCols)]
#
# Auswertegrenzen = data

################################################


# Pfad Windows
Pfad = r"D:\70_Publications\2021\03_Hobeln\50_Experimente\01_Kraftmessungen\Alle\Evaluation"
#Pfad = r"C:\Users\kfabian\polybox\Shared\Share_Kneubühler_Seeholzer\50_Publikationen\05_Critical Feed\02_Kraftmessung_NEW\00_XO_Art391165"


# Pfad Mac:
# Pfad =
# r"/Users/fabiankneubuhler/Desktop/Kraftrauswertung/Python/Daten/Gruppe1"

Dateien = []

# Load all files in directory and sub directories
# r=root, d=directories, f = files
# (https://www.mkyong.com/python/python-how-to-list-all-files-in-a-directory/)
# for r, d, f in os.walk(Pfad):
#     for Datei in f:
#         if ('.dat' in Datei) and ('ProzKraft' in Datei):
#             Dateien.append(os.path.join(r, Datei))

# Load all files in directory
for Datei in os.listdir(Pfad):
    if ('.dat' in Datei) and ('ProzKraft' in Datei):
        Dateien.append(os.path.join(Pfad, Datei))

Auswertegrenzen = Dateien

Experiments = len(Auswertegrenzen) * [0]

# Initialize arrays
# Forces and standard deviations
F_Leerlauf_vorne = np.zeros((len(Auswertegrenzen), 3))
F_Leerlauf_hinten = np.zeros((len(Auswertegrenzen), 3))
Fc = np.zeros((len(Auswertegrenzen), 1))
Ff = np.zeros((len(Auswertegrenzen), 1))
Fp = np.zeros((len(Auswertegrenzen), 1))

# If one array with several experiments

# Define the maximum number of repetitions of experiments in one file
#MaxRepetitions = 1

# AvgFc = np.zeros((len(Auswertegrenzen), MaxRepetitions))
# AvgFf = np.zeros((len(Auswertegrenzen), MaxRepetitions))
# AvgFp = np.zeros((len(Auswertegrenzen), MaxRepetitions))
# SigFc = np.zeros((len(Auswertegrenzen), MaxRepetitions))
# SigFf = np.zeros((len(Auswertegrenzen), MaxRepetitions))
# SigFp = np.zeros((len(Auswertegrenzen), MaxRepetitions))
# # Time points of force evaluation
# TimeForce = np.zeros((len(Auswertegrenzen), (4 * MaxRepetitions)))

# Gradient of linear drift compensation
m_drift = np.zeros((len(Auswertegrenzen), 3))
# Y axis intercept of linear drift compensation
q_drift = np.zeros((len(Auswertegrenzen), 3))
# Time points of linear drift compensation
TimeDrift = np.zeros((len(Auswertegrenzen), 8))

# index_xmin = np.array(0)
# index_xmax = np.array(0)


# Functions

def InitaliseData(Daten):
    Time = np.array(Daten[:, 0])
    # / Wandstaerke = normalisation of forces
    Fc = Daten[:, 1] / Wandstaerke  # Cutting force
    Ff = Daten[:, 3] / Wandstaerke  # Feed force
    Fp = Daten[:, 2] / Wandstaerke  # Passive force

    return Time, Fc, Ff, Fp


def linear_drift_compensation(TimeDrift_start, TimeDrift_end, F_Leerlauf_hinten, F_Leerlauf_vorne):
    m_drift = (F_Leerlauf_hinten - F_Leerlauf_vorne) / \
        (TimeDrift_end - TimeDrift_start)
    q_drift = F_Leerlauf_vorne - m_drift * TimeDrift_start
    return m_drift, q_drift


# Select points
def line_select_callback(eclick, erelease):
    # Returns the selected points via mouse click

    global index_xmax, index_xmin

    x1, y1 = eclick.xdata, eclick.ydata
    x2, y2 = erelease.xdata, erelease.ydata

    mask = (Time > min(x1, x2)) & (Time < max(x1, x2)) & (
        Fc > min(y1, y2)) & (Fc < max(y1, y2))
    xmasked = Time[mask]
    ymasked = Fc[mask]

    if len(ymasked) > 0.0:

        xmax[0] = xmasked.max()
        index_xmaxTest = np.where(Time == xmax[0])
        index_xmax = int(index_xmaxTest[0])
        xmax[1] = Fc[int(index_xmax)]

        xmin[0] = xmasked.min()
        index_xminTest = np.where(Time == xmin[0])
        index_xmin = int(index_xminTest[0])
        xmin[1] = Fc[int(index_xmin)]

        t_max = "xmax: {:.3f}, y(xmax) {:.3f}".format(xmax[0], xmax[1])
        t_min = "xmin: {:.3f}, y(ymin) {:.3f}".format(xmin[0], xmin[1])

        point_min.set_data((xmin[0], xmin[1]))
        point_max.set_data((xmax[0], xmax[1]))

        text_max.set_text(t_max)
        text_max.set_position(((xmax[0] + 0.25), xmax[1]))
        text_min.set_text(t_min)
        text_min.set_position(((xmin[0] - 5), xmin[1]))

        fig.canvas.draw_idle()


def Save_compensation(ZeileExcel, j):
    for ii in range(2, 10):
        EvalDetails.cell(row=ZeileExcel, column=ii,
                         value=TimeDrift[j, (ii - 2)])

    for ii in range(10, 13):
        EvalDetails.cell(row=ZeileExcel, column=ii,
                         value=F_Leerlauf_vorne[j, (ii - 10)])

    for ii in range(13, 16):
        EvalDetails.cell(row=ZeileExcel, column=ii,
                         value=F_Leerlauf_hinten[j, (ii - 13)])

    for ii in range(16, 19):
        EvalDetails.cell(row=ZeileExcel, column=ii,
                         value=m_drift[j, (ii - 16)])

    for ii in range(19, 22):
        EvalDetails.cell(row=ZeileExcel, column=ii,
                         value=q_drift[j, (ii - 19)])

    workbook.save(filename_Excel)
    print("Saved to Excel")


def Save_Forces(ZeileExcel, jj, index_TimeForce, TimeForce):
    MainSheet.cell(row=ZeileExcel, column=jj, value=AvgFc[j, i])
    MainSheet.cell(row=ZeileExcel, column=(jj + 1), value=SigFc[j, i])
    MainSheet.cell(row=ZeileExcel, column=(jj + 2), value=AvgFf[j, i])
    MainSheet.cell(row=ZeileExcel, column=(jj + 3), value=SigFf[j, i])
    MainSheet.cell(row=ZeileExcel, column=(jj + 4), value=AvgFp[j, i])
    MainSheet.cell(row=ZeileExcel, column=(jj + 5), value=SigFp[j, i])

    for ii in range(22, 26):
        EvalDetails.cell(row=ZeileExcel, column=(index_TimeForce + ii),
                         value=TimeForce[j, (index_TimeForce + ii - 22)])
    print("Saved to Excel")

    workbook.save(filename_Excel)


def Load_compensation(ZeileExcel, j):
    #sheet.cell(row=10, column=6).value
    for ii in range(2, 10):
        TimeDrift[j, (ii - 2)] = EvalDetails.cell(row=ZeileExcel,
                                                  column=ii).value

    for ii in range(10, 13):
        F_Leerlauf_vorne[j, (ii - 10)
                         ] = EvalDetails.cell(row=ZeileExcel, column=ii).value

    for ii in range(13, 16):
        F_Leerlauf_hinten[j, (ii - 13)
                          ] = EvalDetails.cell(row=ZeileExcel, column=ii).value

    for ii in range(16, 19):
        m_drift[j, (ii - 16)] = EvalDetails.cell(row=ZeileExcel,
                                                 column=ii).value

    for ii in range(19, 22):
        q_drift[j, (ii - 19)] = EvalDetails.cell(row=ZeileExcel,
                                                 column=ii).value

    return TimeDrift, m_drift, q_drift


def Cut(j, Fc, Ff, Fp, TimeDrift, Time):
    # Cut forces and time signal (drift start to drift end)
    Fc = np.array(Fc[int(TimeDrift[j, 3]):int(TimeDrift[j, 5])])
    # Feed force
    Ff = np.array(Ff[int(TimeDrift[j, 3]):int(TimeDrift[j, 5])])
    # Passive force
    Fp = np.array(Fp[int(TimeDrift[j, 3]):int(TimeDrift[j, 5])])
    Time = np.array(Time[int(TimeDrift[j, 3]):int(TimeDrift[j, 5])])  # Time
    return Fc, Ff, Fp, Time


def Compensate(j, Fc, Ff, Fp, Time, m_drift, q_drift):
    # Drift compensate forces
    Fc = Fc - (Time * m_drift[j, 0] + q_drift[j, 0])
    Ff = Ff - (Time * m_drift[j, 1] + q_drift[j, 1])
    Fp = Fp - (Time * m_drift[j, 2] + q_drift[j, 2])
    return Fc, Ff, Fp


j = -1  # Row index of data arrays (starting at 0)

for Datei in Dateien:
    j = j + 1
    # Read head of "ProzKraft.dat" file (number of experiment, ...)
    D2 = open(Datei, "r")
    Zeile = D2.readline()
    D2.close()

    # Find number of experiment
    VStart = Zeile.find(':') + 1
    VEnde = Zeile.find(' ')
    Versuchsnummer = Zeile[VStart:VEnde]
    CStart = Zeile.find(':', Zeile.find(':') + 1) + 1
    CEnd = Zeile.find('\n')
    Counter = Zeile[CStart:CEnd]
    Experiments[j] = str(Versuchsnummer + '_' + Counter)

    # Find row number of experiment
    for cell in MainSheet['A']:
        # if cell.value == int(Versuchsnummer):
        if cell.value == str(Versuchsnummer + '_' + Counter):
            ZeileExcel = cell.row

    # Find cutting speed
    vc = float(MainSheet.cell(row=ZeileExcel, column=15).value)

    # Find feed
    f = float(MainSheet.cell(row=ZeileExcel, column=12).value)

    # Find Wandstaerke
    Wandstaerke = float(MainSheet.cell(row=ZeileExcel, column=18).value) * 1000

    # Find # force measurements
    Repetitions = int(MainSheet.cell(row=ZeileExcel, column=16).value)
    #Repetitions = 1

    # Initalise array size depending on # Repetions
    AvgFc = np.zeros((len(Auswertegrenzen), Repetitions))
    AvgFf = np.zeros((len(Auswertegrenzen), Repetitions))
    AvgFp = np.zeros((len(Auswertegrenzen), Repetitions))
    SigFc = np.zeros((len(Auswertegrenzen), Repetitions))
    SigFf = np.zeros((len(Auswertegrenzen), Repetitions))
    SigFp = np.zeros((len(Auswertegrenzen), Repetitions))
    # Time points of force evaluation
    TimeForce = np.zeros((len(Auswertegrenzen), (4 * Repetitions)))

    #
    Daten = np.loadtxt(Datei, skiprows=2)

    # Initialise data
    Time, Fc, Ff, Fp = InitaliseData(Daten)

    AnzWerte = Daten[:, 1].size

    #DataToSave = np.column_stack((Experiments[j], AvgFc[j, :], SigFc[j, :], AvgFf[j, :], SigFf[j, :], AvgFp[j, :], SigFp[j, :], [F_Leerlauf_vorne[j, :]]))

    # An area of the force signal can be selected and saved as initial figure
    if (InitialFigure != 0):

        # Idle forces in the front part
        fig, ax = plt.subplots(constrained_layout=True)
        ax.plot(Time, Fc[:], color="blue", linewidth=1.0, linestyle="-",
                label=r'Cutting force [N/mm]')
        ax.plot(Time, Ff[:], color="red", linewidth=1.0, linestyle="-",
                label=r'Feed force [N/mm]')
        ax.plot(Time, Fp[:], color="green",
                linewidth=1.0, linestyle="-", label=r'Passive force [N/mm]')

        # Plot settings
        ax.set_xlabel(r'Time [$s$]')
        ax.set_ylabel(r'Force [N/mm]')
        ax.set_title(r'Raw data of experiment ' + str(Versuchsnummer) + '\_' + str(
            Counter) + ', $v_c=$ ' + str(vc) + ' [m/min]' + ', $f=$ ' + str(f) + ' [mm]')
        ax.legend(loc="lower left")

        print('Select area of interest!')

        # Initiation of interactively selected points
        point_max, = ax.plot([], [], marker="o", color="crimson")
        point_min, = ax.plot([], [], marker="o", color="crimson")
        text_max = ax.text(0, 0, "")
        text_min = ax.text(0, 0, "")

        xmax = [0.0, 0.0]
        xmin = [0.0, 0.0]

        rsa = RectangleSelector(ax, line_select_callback,
                                drawtype='box', useblit=True, button=[1],
                                minspanx=0, minspany=0, spancoords='data',
                                interactive=True)
        plt.show(block=True)
        plt.close()

        # Save selected values as first boundaries of force evaluation
        BeginnMittelung = xmin[0]
        EndeMittelung = xmax[0]

        # Zoomed area of force evaluation
        ZoomBeginn = BeginnMittelung
        ZoomEnde = EndeMittelung
        # Find index of Zoom range
        index_ZoomBeginn = np.where(Time == BeginnMittelung)
        index_ZoomBeginn = int(index_ZoomBeginn[0])
        index_ZoomEnde = np.where(Time == EndeMittelung)
        index_ZoomEnde = int(index_ZoomEnde[0])

        # Plot of zoomed area
        fig, bx = plt.subplots(constrained_layout=True)
        plt.xlim(ZoomBeginn, ZoomEnde)  # Grenzen (min/max) X-Achse
        bx.plot(Time[index_ZoomBeginn:index_ZoomEnde], Fc[index_ZoomBeginn:index_ZoomEnde],
                color="blue", linewidth=1.0, linestyle="-", label=r'Cutting force [N/mm]')
        bx.plot(Time[index_ZoomBeginn:index_ZoomEnde], Ff[index_ZoomBeginn:index_ZoomEnde],
                color="red", linewidth=1.0, linestyle="-", label=r'Feed force [N/mm]')
        bx.plot(Time[index_ZoomBeginn:index_ZoomEnde], Fp[index_ZoomBeginn:index_ZoomEnde],
                color="green", linewidth=1.0, linestyle="-", label=r'Passive force [N/mm]')

        # Plot settings
        bx.legend(loc="lower center")
        plt.xlabel(r'Time [$s$]')
        plt.ylabel(r'Force [N/mm]')
        bx.set_title(r'Cut process forces, experiment ' + str(Versuchsnummer) + "\_" + str(
            Counter) + ', $v_c=$ ' + str(vc) + ' [m/min]' + ', $f=$ ' + str(f) + ' [mm]')

        plt.savefig(r"Figures\EPS\ForceEvaluationCut\ForceEvaluationCut_" +
                    Versuchsnummer + "_" + Counter + ".eps", dpi=300)
        plt.savefig(r"Figures\PNG\ForceEvaluationCut\ForceEvaluationCut_" + Versuchsnummer +
                    "_" + Counter + ".png", dpi=300)
        plt.show(block=True)
        plt.close()

    # Choose interactively forces for drift compensation
    if (DriftCompensationExe != 0):

        # Idle forces in the front part
        fig, cx = plt.subplots(constrained_layout=True)
        cx.plot(Time, Fc[:], color="blue", linewidth=1.0, linestyle="-",
                label=r'Cutting force [N/mm]')
        cx.plot(Time, Ff[:], color="red", linewidth=1.0, linestyle="-",
                label=r'Feed force [N/mm]')
        cx.plot(Time, Fp[:], color="green",
                linewidth=1.0, linestyle="-", label=r'Passive force [N/mm]')

        # Plot settings
        cx.set_xlabel(r'Time [$s$]')
        cx.set_ylabel(r'Force [N/mm]')
        cx.set_title(r'Idle force front, Experiment ' + str(Versuchsnummer) + '\_' + str(
            Counter) + ', $v_c=$ ' + str(vc) + ' [m/min]' + ', $f=$ ' + str(f) + ' [mm]')
        cx.legend(loc="lower left")

        print('Select are for drift compensation -> BEFORE signal rise')

        # Initiation of interactively selected points
        point_max, = cx.plot([], [], marker="o", color="crimson")
        point_min, = cx.plot([], [], marker="o", color="crimson")
        text_max = cx.text(0, 0, "")
        text_min = cx.text(0, 0, "")
        xmax = [0.0, 0.0]
        xmin = [0.0, 0.0]

        rsb = RectangleSelector(cx, line_select_callback,
                                drawtype='box', useblit=True, button=[1],
                                minspanx=0, minspany=0, spancoords='data',
                                interactive=True)
        plt.show(block=True)
        plt.close()

        Stabel = input(('Is the force signal of experiment ' +
                        str(Versuchsnummer) + '_' + str(Counter) + ' stable? [Y/N] '))

        if Stabel == "N" or Stabel == "n":
            MainSheet.cell(row=ZeileExcel, column=26, value='Nein')
            workbook.save(filename_Excel)
            print(
                ('Experiment ' + str(Versuchsnummer) + '_' + str(Counter) + ' is skipped and not evaluated!'))
            continue
        elif Stabel == "y" or Stabel == "y":
            MainSheet.cell(row=ZeileExcel, column=26, value='Ja')
            workbook.save(filename_Excel)
            print("Drift evaluation is continued!")
        else:
            print(
                ('Experiment ' + str(Versuchsnummer) + '_' + str(Counter) + ' is skipped and not evaluated!'))
            continue

        # Save selected values and time as first boundaries of force evaluation

        # Before measurement: Time start point for force averaging
        #TimeDrift[j, 0] = Time[int(index_xmin)]
        TimeDrift[j, 0] = Time[int(index_xmin)]
        # Before measurement: Index start point for force averaging
        TimeDrift[j, 1] = int(index_xmin)
        #TimeDrift[j, 1] = int(index_xmin)
        # Before measurement: Time end point for force averaging
        #TimeDrift[j, 2] = Time[int(index_xmax[0])]
        TimeDrift[j, 2] = Time[int(index_xmax)]
        # Before measurement: Index end point for force averaging
        #TimeDrift[j, 3] = int(index_xmax[0])
        TimeDrift[j, 3] = int(index_xmax)

        F_Leerlauf_vorne[j, 0] = np.mean(
            Fc[int(TimeDrift[j, 1]):int(TimeDrift[j, 3])])
        F_Leerlauf_vorne[j, 1] = np.mean(
            Ff[int(TimeDrift[j, 1]):int(TimeDrift[j, 3])])
        F_Leerlauf_vorne[j, 2] = np.mean(
            Fp[int(TimeDrift[j, 1]):int(TimeDrift[j, 3])])

        # Idle forces in the back part:
        fig, dx = plt.subplots(constrained_layout=True)
        dx.plot(Time, Fc[:], color="blue", linewidth=1.0, linestyle="-",
                label=r'Cutting force [N/mm]')
        dx.plot(Time, Ff[:], color="red", linewidth=1.0, linestyle="-",
                label=r'Feed force [N/mm]')
        dx.plot(Time, Fp[:], color="green",
                linewidth=1.0, linestyle="-", label=r'Passive force [N/mm]')

        # Plot settings
        dx.set_xlabel(r'Time [$s$]')
        dx.set_ylabel(r'Force [N/mm]')
        dx.set_title(r'Idle force back, Experiment ' + str(Versuchsnummer) + "\_" + str(Counter) +
                     ', $v_c=$ ' + str(vc) + ' [m/min]' + ', $f=$ ' + str(f) + ' [mm]')
        dx.legend(loc="lower left")

        print('Select are for drift compensation -> AFTER signal fall')

        # Initiation of interactively selected points
        point_max, = dx.plot([], [], marker="o", color="crimson")
        point_min, = dx.plot([], [], marker="o", color="crimson")
        text_max = dx.text(0, 0, "")
        text_min = dx.text(0, 0, "")
        xmax = [0.0, 0.0]
        xmin = [0.0, 0.0]

        rsc = RectangleSelector(dx, line_select_callback,
                                drawtype='box', useblit=True, button=[1],
                                minspanx=0, minspany=0, spancoords='data',
                                interactive=True)
        plt.show(block=True)
        plt.close()

        # Save selected values and time as second boundaries of force
        # evaluation
        # After measurement: Time start point for force averaging
        TimeDrift[j, 4] = Time[int(index_xmin)]
        # After measurement: Index start point for force averaging
        TimeDrift[j, 5] = int(index_xmin)
        # After measurement: Time end point for force averaging
        TimeDrift[j, 6] = Time[int(index_xmax)]
        # After measurement: Index end point for force averaging
        TimeDrift[j, 7] = int(index_xmax)

        F_Leerlauf_hinten[j, 0] = np.mean(
            Fc[int(TimeDrift[j, 5]):int(TimeDrift[j, 7])])
        F_Leerlauf_hinten[j, 1] = np.mean(
            Ff[int(TimeDrift[j, 5]):int(TimeDrift[j, 7])])
        F_Leerlauf_hinten[j, 2] = np.mean(
            Fp[int(TimeDrift[j, 5]):int(TimeDrift[j, 7])])

        # Linear compensation function
        m_drift, q_drift = linear_drift_compensation(
            TimeDrift[j, 2], TimeDrift[j, 5], F_Leerlauf_hinten, F_Leerlauf_vorne)

        # Save to Excel
        if EvalDetails.cell(row=ZeileExcel, column=2).value == None:
            Save_compensation(ZeileExcel, j)

        else:
            Override = input(
                ('Do you want to override the drift compensation of experiment ' + str(Versuchsnummer) + "_" + str(Counter) + '? [Y/N] '))
            if Override == "N" or Override == "n":
                TimeDrift, m_drift, q_drift = Load_compensation(ZeileExcel, j)
            else:
                Save_compensation(ZeileExcel, j)

        Fc, Ff, Fp, Time = Cut(j, Fc, Ff, Fp, TimeDrift, Time)

        #input("Press Enter to continue!")

        Fc_uncompensated = Fc
        Ff_uncompensated = Ff
        Fp_uncompensated = Fp

        Fc, Ff, Fp = Compensate(j, Fc, Ff, Fp, Time, m_drift, q_drift)

        # Plot of drift compensation
        fig, ex = plt.subplots(constrained_layout=True)
        ex.plot(Time, Fc, color="blue", linewidth=1.0,
                linestyle="-", label=r'Cutting force [N/mm]')
        ex.plot(Time, Fc_uncompensated, color="blue", linewidth=1.0,
                linestyle="--", label=r'Cutting force (uncompensated) [N/mm]')
        ex.plot(Time, Ff, color="red", linewidth=1.0,
                linestyle="-", label=r'Feed force [N/mm]')
        ex.plot(Time, Ff_uncompensated, color="red", linewidth=1.0,
                linestyle="--", label=r'Feed force (uncompensated) [N/mm]')
        ex.plot(Time, Fp, color="green", linewidth=1.0,
                linestyle="-", label=r'Passive force [N/mm]')
        ex.plot(Time, Fp_uncompensated, color="green", linewidth=1.0,
                linestyle="--", label=r'Passive force (uncompensated) [N/mm]')

        ex.plot(Time, (Time * m_drift[j, 0] + q_drift[j, 0]), color="blue",
                linewidth=2.0, linestyle="--", label=r'Drift')
        ex.plot(Time[:], (Time[:] * m_drift[j, 1] + q_drift[j, 1]),
                color="red", linewidth=2.0, linestyle="--", label=r'Drift')
        ex.plot(Time[:], (Time[:] * m_drift[j, 2] + q_drift[j, 2]),
                color="green", linewidth=2.0, linestyle="--", label=r'Drift')

        # Plot settings
        ex.set_xlabel(r'Time [$s$]')
        ex.set_ylabel(r'Force [N/mm]')
        ex.set_title(r'Drift compensation: Experiment ' + str(Versuchsnummer) + "\_" + str(Counter) +
                     ', $v_c=$ ' + str(vc) + ' [m/min]' + ', $f=$ ' + str(f) + ' [mm]')
        ex.legend(loc="lower center")

        plt.savefig(r"Figures\EPS\DriftCompensation\DriftCompensation_" +
                    Versuchsnummer + "_" + Counter + ".eps", dpi=300)
        plt.savefig(r"Figures\PNG\DriftCompensation\DriftCompensation_" +
                    Versuchsnummer + "_" + Counter + ".png", dpi=300)

        plt.show(block=True)
        plt.close()

    for i in range(int(Repetitions)):

        if not EvalDetails.cell(row=ZeileExcel, column=2).value == None and DriftCompensation != 0 and not DriftCompensationExe == 1:
            Time, Fc, Ff, Fp = InitaliseData(Daten)
            TimeDrift, m_drift, q_drift = Load_compensation(ZeileExcel, j)
            Fc, Ff, Fp, Time = Cut(j, Fc, Ff, Fp, TimeDrift, Time)
            Fc, Ff, Fp = Compensate(j, Fc, Ff, Fp, Time, m_drift, q_drift)

        elif EvalDetails.cell(row=ZeileExcel, column=2).value == None and DriftCompensation != 0:
            NoDrift = input(('No drift compensation available for' +
                             str(Versuchsnummer) + "_" + str(Counter) + ' experiment! Please confirm with [Y] '))
            if NoDrift == "Y" or NoDrift == "y":
                print("Thank you!")
            else:
                print(
                    ('Experiment ' + str(Versuchsnummer) + '\_' + str(Counter) + ' is not evaluated!'))
                continue

        # Choose interactively borders for force evaluation
        fig, ax = plt.subplots(constrained_layout=True)
        plt.xlim(TimeDrift[j, 2], TimeDrift[j, 4])
        ax.plot(Time, Fc, color="blue", linewidth=1.0, linestyle="-",
                label=r'Cutting force [N/mm]')  # Korrigierte Zuordnung der Prozeßkräfte, HK, Fr, 05.06.2020
        ax.plot(Time, Ff, color="red", linewidth=1.0, linestyle="-",
                label=r'Feed force [N/mm]')  # Korrigierte Zuordnung der Prozeßkräfte, HK, Fr, 05.06.2020
        ax.plot(Time, Fp, color="green",
                linewidth=1.0, linestyle="-", label=r'Passive force [N/mm]')

        # Plot settings
        ax.set_xlabel(r'Time [$s$]')
        ax.set_ylabel(r'Force [N/mm]')
        ax.set_title(r'Process forces, Experiment ' + str(Versuchsnummer) + "\_" + str(Counter) +
                     ', $v_c=$ ' + str(vc) + ' [m/min]' + ', $f=$ ' + str(f) + ' [mm]')
        ax.legend(loc="lower center")

        print('Select force measurement no.' + str(i + 1) +
              ' OR close the window twice if the signal is not stable! Start from the LEFT side (=1) and continue to the RIGHT side.')

        # Initiation of interactively selected points
        point_max, = ax.plot([], [], marker="o", color="crimson")
        point_min, = ax.plot([], [], marker="o", color="crimson")
        text_max = ax.text(TimeDrift[j, 2], 0, "")
        text_min = ax.text(TimeDrift[j, 2], 0, "")
        xmax = [TimeDrift[j, 2], 0.0]
        xmin = [TimeDrift[j, 2], 0.0]

        # Choose rough area of force evaluation
        rs = RectangleSelector(ax, line_select_callback,
                               drawtype='box', useblit=True, button=[1],
                               minspanx=0, minspany=0, spancoords='data',
                               interactive=True)

        plt.savefig(r"Figures\EPS\ForceEvaluationCut\ForceEvaluationCutCompensated_" +
                    Versuchsnummer + "_" + Counter + ".eps", dpi=300)
        plt.savefig(r"Figures\PNG\ForceEvaluationCut\ForceEvaluationCutCompensated_" +
                    Versuchsnummer + "_" + Counter + ".png", dpi=300)

        plt.show(block=True)
        plt.close()

        # Save selected values as first boundaries of force evaluation
        BeginnMittelung = xmin[0]
        EndeMittelung = xmax[0]

        # Zoomed area of force evaluation
        ZoomBeginn = BeginnMittelung
        ZoomEnde = EndeMittelung
        # Find index of Zoom range
        index_ZoomBeginn = np.where(Time == BeginnMittelung)
        index_ZoomBeginn = int(index_ZoomBeginn[0])
        index_ZoomEnde = np.where(Time == EndeMittelung)
        index_ZoomEnde = int(index_ZoomEnde[0])

        # Plot of zoomed area
        fig, cx = plt.subplots(constrained_layout=True)
        plt.xlim(ZoomBeginn, ZoomEnde)  # Grenzen (min/max) X-Achse
        cx.plot(Time[index_ZoomBeginn:index_ZoomEnde], Fc[index_ZoomBeginn:index_ZoomEnde],
                color="blue", linewidth=1.0, linestyle="-", label=r'Cutting force [N/mm]')
        cx.plot(Time[index_ZoomBeginn:index_ZoomEnde], Ff[index_ZoomBeginn:index_ZoomEnde],
                color="red", linewidth=1.0, linestyle="-", label=r'Feed force [N/mm]')
        cx.plot(Time[index_ZoomBeginn:index_ZoomEnde], Fp[index_ZoomBeginn:index_ZoomEnde],
                color="green", linewidth=1.0, linestyle="-", label=r'Passive force [N/mm]')

        # Plot settings
        cx.legend(loc="lower left")
        plt.xlabel(r'Time [$s$]')
        plt.ylabel(r'Force [N/mm]')
        cx.set_title(r'Process forces, Experiment ' + str(Versuchsnummer) + "\_" + str(Counter) +
                     "\_" + str(i + 1) + ', $v_c=$ ' + str(vc) + ' [m/min]' + ', $f=$ ' + str(f) + ' [mm]')

        # Initiation of interactively selected data
        point_max, = cx.plot([], [], marker="o", color="crimson")
        point_min, = cx.plot([], [], marker="o", color="crimson")
        line_FcAvg, = cx.plot([], [], color="darkblue", linestyle='--')
        line_FfAvg, = cx.plot([], [], color="darkred", linestyle='--')
        line_FpAvg, = cx.plot([], [], color="darkgreen", linestyle='--')

        text_max = cx.text(0, 0, "")
        text_min = cx.text(0, 0, "")
        text_AvgFc = cx.text(0, 0, "")
        text_SigFc = cx.text(0, 0, "")
        text_AvgFf = cx.text(0, 0, "")
        text_SigFf = cx.text(0, 0, "")
        text_AvgFp = cx.text(0, 0, "")
        text_SigFp = cx.text(0, 0, "")

        def line_select_callback_force(eclick, erelease):
            x1, y1 = eclick.xdata, eclick.ydata
            x2, y2 = erelease.xdata, erelease.ydata

            # Select x values in rectangle
            #mask_x = (Time > min(x1, x2)) & (Time < max(x1, x2))
            #xmasked = Time[mask_x]

            mask_x = (Time > min(x1, x2)) & (Time < max(x1, x2)) & (
                Fc > min(y1, y2)) & (Fc < max(y1, y2))
            xmasked = Time[mask_x]

            if len(xmasked) > 0.0:

                global index_xmin_force
                global index_xmax_force

                xmax[0] = xmasked.max()
                index_xmax_force = np.where(Time == xmax[0])
                xmax[1] = Fc[int(index_xmax_force[0])]

                xmin[0] = xmasked.min()
                index_xmin_force = np.where(Time == xmin[0])
                xmin[1] = Fc[int(index_xmin_force[0])]

                # Calculate average force and standard deviation
                AvgFc[j, i] = np.round(np.mean(
                    Fc[int(index_xmin_force[0]):int(index_xmax_force[0])]), 1)
                SigFc[j, i] = np.round(np.std(
                    Fc[int(index_xmin_force[0]):int(index_xmax_force[0])]), 2)
                AvgFf[j, i] = np.round(np.mean(
                    Ff[int(index_xmin_force[0]):int(index_xmax_force[0])]), 1)
                SigFf[j, i] = np.round(np.std(
                    Ff[int(index_xmin_force[0]):int(index_xmax_force[0])]), 2)
                AvgFp[j, i] = np.round(np.mean(
                    Fp[int(index_xmin_force[0]):int(index_xmax_force[0])]), 1)
                SigFp[j, i] = np.round(np.std(
                    Fp[int(index_xmin_force[0]):int(index_xmax_force[0])]), 2)

                # Setup text on plot
                #t_max = "xmax: {:.3f}, y(xmax) {:.3f}".format(xmax[0], xmax[1])
                #t_min = "xmin: {:.3f}, y(ymin) {:.3f}".format(xmin[0], xmin[1])
                t_AvgFc = "Avgerage cutting force: {:.1f} N, SD {:.2f} N".format(
                    AvgFc[j, i], SigFc[j, i])
                t_AvgFf = "Average feed force: {:.1f} N, SD {:.2f} N".format(
                    AvgFf[j, i], SigFf[j, i])
                t_AvgFp = "Average passive force: {:.1f} N, SD {:.2f} N".format(
                    AvgFp[j, i], SigFp[j, i])

                point_min.set_data((xmin[0], xmin[1]))
                point_max.set_data((xmax[0], xmax[1]))
                line_FcAvg.set_data(
                    Time[index_ZoomBeginn:index_ZoomEnde], AvgFc[j, i])
                line_FfAvg.set_data(
                    Time[index_ZoomBeginn:index_ZoomEnde], AvgFf[j, i])
                line_FpAvg.set_data(
                    Time[index_ZoomBeginn:index_ZoomEnde], AvgFp[j, i])

                text_AvgFc.set_text(t_AvgFc)
                text_AvgFc.set_position(
                    ((((xmax[0]) - xmin[0]) / 5 + xmin[0]), AvgFc[j, i] + 2 * SigFc[j, i] + 1))
                text_AvgFc.set_fontsize(14)

                text_AvgFf.set_text(t_AvgFf)
                text_AvgFf.set_position(
                    ((((xmax[0]) - xmin[0]) / 5 + xmin[0]), AvgFf[j, i] + 2 * SigFf[j, i] + 1))
                text_AvgFf.set_fontsize(14)

                text_AvgFp.set_text(t_AvgFp)
                text_AvgFp.set_position(
                    ((((xmax[0]) - xmin[0]) / 5 + xmin[0]), AvgFp[j, i] + 2 * SigFp[j, i] + 1))
                text_AvgFp.set_fontsize(14)

                # X max and Y min points description
                # text_max.set_text(t_max)
                # text_max.set_position(((xmax[0]), xmax[1]))
                # text_min.set_text(t_min)
                # text_min.set_position(((xmin[0]), xmin[1]))

                fig.canvas.draw_idle()

                plt.savefig(r"Figures\EPS\SingleForceEvaluations\ForceEvaluation_" + Versuchsnummer +
                            "_" + Counter + "_" + str(i + 1) + ".eps", dpi=300)
                plt.savefig(r"Figures\PNG\SingleForceEvaluations\ForceEvaluation_" + Versuchsnummer +
                            "_" + Counter + "_" + str(i + 1) + ".png", dpi=300)

        rs = RectangleSelector(cx, line_select_callback_force,
                               drawtype='box', useblit=True, button=[1],
                               minspanx=0, minspany=0, spancoords='pixels',
                               interactive=True)

        plt.show(block=True)
        plt.close()

        # Is signal stable?
        Stabel = input(('Is the force signal of experiment ' +
                        str(Versuchsnummer) + '_' + str(Counter) + '_' + str(i + 1) + ' stable? [Y/N] '))

        if Stabel == "N" or Stabel == "n":
            # Average force and standard deviation = 0
            AvgFc[j, i] = 0.0
            SigFc[j, i] = 0.0
            AvgFf[j, i] = 0.0
            SigFf[j, i] = 0.0
            AvgFp[j, i] = 0.0
            SigFp[j, i] = 0.0

            print(
                ('Experiment ' + str(Versuchsnummer) + '_' + str(Counter) + '_' + str(i + 1) + ' is skipped and not evaluated!'))
            continue
        elif Stabel == "y" or Stabel == "y":
            print("Force evaluation is continued!")
        else:
            print(
                ('Experiment ' + str(Versuchsnummer) + '_' + str(Counter) + '_' + str(i + 1) + ' is skipped and not evaluated!'))
            continue

        # Save selected values and time of force evaluation
        # Time start point for force averaging

        index_TimeForce = int(i * 4)

        TimeForce[j, index_TimeForce] = Time[int(index_xmin_force[0])]
        # Index start point for force averaging
        TimeForce[j, (index_TimeForce + 1)] = int(index_xmin_force[0])
        # Time end point for force averaging
        TimeForce[j, (index_TimeForce + 2)] = Time[int(index_xmax_force[0])]
        # Index end point for force averaging
        TimeForce[j, (index_TimeForce + 3)] = int(index_xmax_force[0])

        # Save data to Excel
        jj = i * 6 + 27

        if MainSheet.cell(row=ZeileExcel, column=jj).value == None:
            Save_Forces(ZeileExcel, jj, index_TimeForce, TimeForce)

        else:
            Override = input(
                ('Do you want to override the force values of experiment ' + str(Versuchsnummer) + "_" + str(Counter) + ' repetition ' + str(i) + '? [y/n] '))
            if Override == "N" or Override == "n":
                continue
            elif Override == "Y" or "y":
                Save_Forces(ZeileExcel, jj, index_TimeForce, TimeForce)

    Head = str('Experiment, Cutting force F_c [N/mm], SD F_c [N/mm], Feed force F_f [N/mm], SD F_f [N/mm], Passive force F_p [N/mm], SD F_p [N/mm], TimeFront1, IndTimeFront1, TimeFront2, IndTimeFront2, TimeBack1, IndTimeBack1, TimeBack2, IndTimeBack2, F_Leerlauf_vorne[0], F_Leerlauf_vorne[1], F_Leerlauf_vorne[2], F_Leerlauf_hinten[0], F_Leerlauf_hinten[1], F_Leerlauf_hinten[2], m_drift[0], m_drift[1], m_drift[2], q_drift[0], q_drift[1], q_drift[2], TimeForce')
    DataToSave = np.column_stack((Experiments[j], [AvgFc[j, :]], [SigFc[j, :]], [AvgFf[j, :]], [SigFf[j, :]], [AvgFp[j, :]], [SigFp[j, :]], [
                                 F_Leerlauf_vorne[j, :]], [F_Leerlauf_hinten[j, :]], [m_drift[j, :]], [q_drift[j, :]], [TimeDrift[j, :]], [TimeForce[j, :]]))
    Datafile_path = r"Evaluation_files\ProcessForceEvaluation_" + \
        Versuchsnummer + "_" + Counter + ".txt"
    np.savetxt(Datafile_path, DataToSave, fmt='%s', delimiter=' ', header=Head)

    print("Text file of experiment " +
          Versuchsnummer + "_" + Counter + " saved!")


# numpy.savetxt(file_name, Array, format='%.4e', delimiter='  ',
# newline='n', header='  ', footer='  ', comments='# ')

# Save entire arrays to 1 text file

#Head = str('Experiment, Cutting force F_c [N/mm], SD F_c [N/mm], Feed force F_f [N/mm], SD F_f [N/mm], Passive force F_p [N/mm], SD F_p [N/mm], TimeFront1, IndTimeFront1, TimeFront2, IndTimeFront2, TimeBack1, IndTimeBack1, TimeBack2, IndTimeBack2, F_Leerlauf_vorne[0], F_Leerlauf_vorne[1], F_Leerlauf_vorne[2], F_Leerlauf_hinten[0], F_Leerlauf_hinten[1], F_Leerlauf_hinten[2], m_drift[0], m_drift[1], m_drift[2], q_drift[0], q_drift[1], q_drift[2], TimeForce[0], TimeForce[1, TimeForce[2], TimeForce[3], TimeForce[4], TimeForce[5], TimeForce[6], TimeForce[7], TimeForce[8], TimeForce[9], TimeForce[10], TimeForce[11], TimeForce[12], TimeForce[13], TimeForce[14], TimeForce[15]')
# Head = str('Experiment, Cutting force F_c [N/mm], SD F_c [N/mm], Feed force F_f [N/mm], SD F_f [N/mm], Passive force F_p [N/mm], SD F_p [N/mm], TimeFront1, IndTimeFront1, TimeFront2, IndTimeFront2, TimeBack1, IndTimeBack1, TimeBack2, IndTimeBack2, F_Leerlauf_vorne[0], F_Leerlauf_vorne[1], F_Leerlauf_vorne[2], F_Leerlauf_hinten[0], F_Leerlauf_hinten[1], F_Leerlauf_hinten[2], m_drift[0], m_drift[1], m_drift[2], q_drift[0], q_drift[1], q_drift[2], TimeForce')
# DataToSave = np.column_stack(
#     [Experiments, AvgFc, SigFc, AvgFf, SigFf, AvgFp, SigFp, F_Leerlauf_vorne, F_Leerlauf_hinten, m_drift, q_drift, TimeDrift, TimeForce])
# Datafile_path = r"ProcessForceEvaluation_" + \
#     Experiments[0] + "_" + Experiments[-1] + ".txt"
# np.savetxt(Datafile_path, DataToSave, fmt='%s', delimiter=' ', header=Head)


# ProzKraftAusgabe = "ProcessForceEvaluation.txt"
#
# Datei = open(ProzKraftAusgabe, 'w')
# Datei.write('Experiment, Cutting force F_c [N/mm], SD F_c [N], Feed force F_f [N/mm], SD F_f [N], Passive force F_p [N/mm], SD F_p [N], TimeFront1, IndTimeFront1, TimeFront2, IndTimeFront2, TimeBack1, IndTimeBack1, TimeBack2, IndTimeBack2, F_Leerlauf_vorne[0], F_Leerlauf_vorne[1], F_Leerlauf_vorne[2], F_Leerlauf_hinten[0], F_Leerlauf_hinten[1], F_Leerlauf_hinten[2], m_drift[0], m_drift[1], m_drift[2], q_drift[0], q_drift[1], q_drift[2], TimeForce[0], TimeForce[1, TimeForce[2], TimeForce[3], TimeForce[4], TimeForce[5], TimeForce[6], TimeForce[7], TimeForce[8], TimeForce[9], TimeForce[10], TimeForce[11], TimeForce[12], TimeForce[13], TimeForce[14], TimeForce[15]\n')
# #
# for i in range(len(Experiments)):
#     Datei.write(str(Experiments[i]) + ', ')
#     for ii in range(Repetitions):
#         Datei.write(str(AvgFc[i, ii]) + ', ' + str(SigFc[i, ii]) + ', ' + str(AvgFf[i, ii]) +
#                     ', ' + str(SigFf[i, ii]) + ', ' + str(AvgFp[i, ii]) + ', ' + str(SigFp[i, ii]) + ', ')
#     for ii in range(3):
#         Datei.write(str(F_Leerlauf_vorne[i, ii]) +
#                     ', ' + str(F_Leerlauf_vorne[i, ii]) + ', ')
#     for ii in range(2):
#         Datei.write(str(m_drift[i, ii]) + ', ' + str(q_drift[i, ii]) + ', ')
#     for ii in range(8):
#         Datei.write(str(TimeDrift[i, ii]) + ', ')
#     for ii in range((Repetitions * 2)):
#         Datei.write(str(TimeForce[i, ii]) + ', ')
#     Datei.write('\n')
#
# Datei.close()

#print("Text file saved")
