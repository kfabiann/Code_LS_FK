# Python and Excel
# Author: Fabian Kneubuehler

# 24.08.2021


#-*- coding:utf-8 -*-


####################################################
####################################################
# Create a Excel file with content:

# from openpyxl import Workbook
#
# workbook = Workbook()
# sheet = workbook.active
#
# sheet["A1"] = "hello"
# sheet["B1"] = "world!"
#
# workbook.save(filename="hello_world.xlsx")


# Save to Excel

# for i in range(5,15):
#     cellref=EvalDetails.cell(row=ZeileExcel, column=i)
#     cellref.value=lista[i]

####################################################
####################################################

# Plot something
import matplotlib.pyplot as plt
import pickle

# fig, ax = plt.subplots()
# ax.plot([1, 2, 3], [10, -10, 30])
#
# pickle.dump(fig, open('FigureObject.fig.pickle', 'wb'))

figx = pickle.load(open('FigureObject.fig.pickle', 'rb'))


####################################################
####################################################


# Read data from Excel file


def goto(linenum):
    global line
    line = linenum


line = 1
while True:
    if line == 1:
        response = input("yes or no? ")
        if response == "yes":
            goto(2)
        elif response == "no":
            goto(3)
        else:
            goto(100)
    elif line == 2:
        print("Thank you for the yes!")
        goto(20)
    elif line == 3:
        print("Thank you for the no!")
        goto(20)
    elif line == 20:
        break
    elif line == 100:
        print("You're annoying me - answer the question!")
        goto(1)


from openpyxl import load_workbook
import numpy as np
from array import array

workbook = load_workbook(
    filename="Zerspanungsversuche_Hobeln_Protokoll_Gesamt_20082021.xlsx", data_only=True)
workbook.sheetnames

sheet = workbook.active
a = sheet["A8"].value

####################################################
# Import data from Excel file

# for value in sheet.iter_rows(min_row=8, max_row=9, min_col=1, max_col=10, values_only=True):
#     print(value)


# read one specific row from excel
firstRow = 8
firstCol = 1
nCols = 30
nRows = 12


allCells = np.array([[cell.value for cell in row]
                     for row in sheet.iter_rows()])


data = allCells[(firstRow - 1):(firstRow - 1 + nRows),
                (firstCol - 1):(firstCol - 1 + nCols)]
# print(data)

# print(data)
#print(data[:, 0])

bb = np.where(523 == data[:, 0])

print(bb[0])

test = bb[0] * bb[0]
print(int(test))

# Konfigurierte Zeile: Versuch No, Schnittgeschw., Vorschubgeschw.,
# Start1, Ende1, ...
Zeile_konfig = [data[0][0], data[0][14], data[0][11], data[0][26], data[0][27]]

print(Zeile_konfig)


# for column in sheet.iter_cols(min_row=8,
#                               max_row=10,
#                               min_col=0,
#                               max_col=3,
#                               values_only=True):
#     print(column)


# for row in sheet.rows:
#     print(row)
