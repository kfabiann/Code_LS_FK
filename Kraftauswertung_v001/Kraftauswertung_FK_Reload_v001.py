"""
Script based on: "Auswertung_Prozeszkraefte_Zerspanungsversuche.py" @Hagen Klippel

Modified: 25.08.2021
@author: Fabian Kneubuehler

Theme: Evaluation of process forces 

#-*- coding:utf-8 -*-
"""

import numpy as np
import matplotlib.pyplot as plt
import os
import openpyxl
from openpyxl import load_workbook

# https://matplotlib.org/users/usetex.html
plt.rc('text', usetex=True)
plt.rc('font', family='serif')

# Drift compensation
DriftCompensation = 1  # Yes=1, No=0


# Load Excel file
filename_Excel = "Zerspanungsversuche_Hobeln_Protokoll_Gesamt_03102021.xlsx"
workbook = load_workbook(
    filename=filename_Excel, data_only=True)

#sheet = workbook.active
#sheet = workbook.get_sheet_by_name('Overview')
MainSheet = workbook.worksheets[0]
EvalDetails = workbook.worksheets[1]

# Pfad Windows
Pfad = r"D:\70_Publications\2021\03_Hobeln\50_Experimente\01_Kraftmessungen\Alle\Evaluation"
Path_eval_files = r"C:\Users\kfabian\eclipse-workspace\Kraftauswertung_v001\Evaluation_files"

# Pfad Mac:
# Pfad =
# r"/Users/fabiankneubuhler/Desktop/Kraftrauswertung/Python/Daten/Gruppe1"

Dateien = []

# %%

# Load dat file
for Datei in os.listdir(Pfad):
    if ('.dat' in Datei) and ('ProzKraft' in Datei):
        Dateien.append(os.path.join(Pfad, Datei))

Auswertegrenzen = Dateien
#Auswertegrenzen = [0]

Experiments = len(Auswertegrenzen) * [0]


##########################################################################
##########################################################################
# Initialize arrays

# Forces and standard deviations
F_Leerlauf_vorne = np.zeros((len(Auswertegrenzen), 3))
F_Leerlauf_hinten = np.zeros((len(Auswertegrenzen), 3))
Fc = np.zeros((len(Auswertegrenzen), 1))
Ff = np.zeros((len(Auswertegrenzen), 1))
Fp = np.zeros((len(Auswertegrenzen), 1))

# Gradient of linear drift compensation
m_drift = np.zeros((len(Auswertegrenzen), 3))
# Y axis intercept of linear drift compensation
q_drift = np.zeros((len(Auswertegrenzen), 3))
# Time points of linear drift compensation
TimeDrift = np.zeros((len(Auswertegrenzen), 8))


##########################################################################
##########################################################################
# Functions


def InitaliseData(Daten):
    Time = np.array(Daten[:, 0])
    # / Wandstaerke = normalisation of forces
    Fc = Daten[:, 1] / Wandstaerke  # Cutting force
    Ff = Daten[:, 3] / Wandstaerke  # Feed force
    Fp = Daten[:, 2] / Wandstaerke  # Passive force

    return Time, Fc, Ff, Fp


def Save_compensation(ZeileExcel, j):
    for ii in range(2, 10):
        EvalDetails.cell(row=ZeileExcel, column=ii,
                         value=TimeDrift[j, (ii - 2)])

    for ii in range(10, 13):
        EvalDetails.cell(row=ZeileExcel, column=ii,
                         value=F_Leerlauf_vorne[j, (ii - 10)])

    for ii in range(13, 16):
        EvalDetails.cell(row=ZeileExcel, column=ii,
                         value=F_Leerlauf_hinten[j, (ii - 13)])

    for ii in range(16, 19):
        EvalDetails.cell(row=ZeileExcel, column=ii,
                         value=m_drift[j, (ii - 16)])

    for ii in range(19, 22):
        EvalDetails.cell(row=ZeileExcel, column=ii,
                         value=q_drift[j, (ii - 19)])

    workbook.save(filename_Excel)
    print("Saved to Excel")


def Save_Forces(ZeileExcel, jj, index_TimeForce, TimeForce):
    MainSheet.cell(row=ZeileExcel, column=jj, value=AvgFc[j, i])
    MainSheet.cell(row=ZeileExcel, column=(jj + 1), value=SigFc[j, i])
    MainSheet.cell(row=ZeileExcel, column=(jj + 2), value=AvgFf[j, i])
    MainSheet.cell(row=ZeileExcel, column=(jj + 3), value=SigFf[j, i])
    MainSheet.cell(row=ZeileExcel, column=(jj + 4), value=AvgFp[j, i])
    MainSheet.cell(row=ZeileExcel, column=(jj + 5), value=SigFp[j, i])

    for ii in range(22, 26):
        EvalDetails.cell(row=ZeileExcel, column=(index_TimeForce + ii),
                         value=TimeForce[j, (index_TimeForce + ii - 22)])
    print("Saved to Excel")

    workbook.save(filename_Excel)


def Load_compensation(ZeileExcel, j):
    #sheet.cell(row=10, column=6).value
    for ii in range(2, 10):
        TimeDrift[j, (ii - 2)] = EvalDetails.cell(row=ZeileExcel,
                                                  column=ii).value

    for ii in range(10, 13):
        F_Leerlauf_vorne[j, (ii - 10)
                         ] = EvalDetails.cell(row=ZeileExcel, column=ii).value

    for ii in range(13, 16):
        F_Leerlauf_hinten[j, (ii - 13)
                          ] = EvalDetails.cell(row=ZeileExcel, column=ii).value

    for ii in range(16, 19):
        m_drift[j, (ii - 16)] = EvalDetails.cell(row=ZeileExcel,
                                                 column=ii).value

    for ii in range(19, 22):
        q_drift[j, (ii - 19)] = EvalDetails.cell(row=ZeileExcel,
                                                 column=ii).value

    return TimeDrift, m_drift, q_drift


def Load_all_data_txt(Data_row, Repetitions):
    for i in range(Repetitions):
        AvgFc[0, i] = Data_row[i + 1]
    for i in range(Repetitions):
        SigFc[0, i] = Data_row[i + Repetitions + 1]
    for i in range(Repetitions):
        AvgFf[0, i] = Data_row[i + 2 * Repetitions + 1]
    for i in range(Repetitions):
        SigFf[0, i] = Data_row[i + 3 * Repetitions + 1]
    for i in range(Repetitions):
        SigFp[0, i] = Data_row[i + 4 * Repetitions + 1]
    for i in range(Repetitions):
        SigFp[0, i] = Data_row[i + 5 * Repetitions + 1]
    for i in range(3):
        F_Leerlauf_vorne[0, i] = Data_row[i + 6 * Repetitions + 1]
    for i in range(3):
        F_Leerlauf_hinten[0, i] = Data_row[i + 6 * Repetitions + 4]
    for i in range(3):
        m_drift[0, i] = Data_row[i + 6 * Repetitions + 7]
    for i in range(3):
        q_drift[0, i] = Data_row[i + 6 * Repetitions + 10]
    for i in range(8):
        TimeDrift[0, i] = Data_row[i + 6 * Repetitions + 13]
    for i in range(4 * Repetitions):
        TimeForce[0, i] = Data_row[i + 6 * Repetitions + 21]

    return AvgFc, SigFc, AvgFf, SigFf, AvgFp, SigFp, F_Leerlauf_vorne, F_Leerlauf_hinten, m_drift, q_drift, TimeDrift, TimeForce


def Load_compensation_txt(Data_row, Repetitions):

    for i in range(3):
        m_drift[0, i] = Data_row[i + 6 * Repetitions + 7]
    for i in range(3):
        q_drift[0, i] = Data_row[i + 6 * Repetitions + 10]
    for i in range(8):
        TimeDrift[0, i] = Data_row[i + 6 * Repetitions + 13]

    return m_drift, q_drift, TimeDrift


def Cut(Fc, Ff, Fp, TimeDrift, Time):
    # Cut forces and time signal (drift start to drift end)
    Fc = np.array(Fc[int(TimeDrift[0, 3]):int(TimeDrift[0, 5])])
    # Feed force
    Ff = np.array(Ff[int(TimeDrift[0, 3]):int(TimeDrift[0, 5])])
    # Passive force
    Fp = np.array(Fp[int(TimeDrift[0, 3]):int(TimeDrift[0, 5])])
    Time = np.array(Time[int(TimeDrift[0, 3]):int(TimeDrift[0, 5])])  # Time
    return Fc, Ff, Fp, Time


def Compensate(Fc, Ff, Fp, Time, m_drift, q_drift):
    # Drift compensate forces
    Fc = Fc - (Time * m_drift[0, 0] + q_drift[0, 0])
    Ff = Ff - (Time * m_drift[0, 1] + q_drift[0, 1])
    Fp = Fp - (Time * m_drift[0, 2] + q_drift[0, 2])
    return Fc, Ff, Fp


def Load_ForceEval_txt(Data_row, Repetitions):
    #TimeDrift, m_drift, q_drift
    for i in range(Repetitions):
        AvgFc[0, i] = Data_row[i + 1]
    for i in range(Repetitions):
        SigFc[0, i] = Data_row[i + Repetitions + 1]
    for i in range(Repetitions):
        AvgFf[0, i] = Data_row[i + 2 * Repetitions + 1]
    for i in range(Repetitions):
        SigFf[0, i] = Data_row[i + 3 * Repetitions + 1]
    for i in range(Repetitions):
        SigFp[0, i] = Data_row[i + 4 * Repetitions + 1]
    for i in range(Repetitions):
        SigFp[0, i] = Data_row[i + 5 * Repetitions + 1]
    for i in range(4 * Repetitions):
        TimeForce[0, i] = Data_row[i + 6 * Repetitions + 21]

    return AvgFc, SigFc, AvgFf, SigFf, AvgFp, SigFp, TimeForce


##########################################################################
##########################################################################

j = -1  # Row index of data arrays (starting at 0)

for Datei in Dateien:
    j = j + 1
    # Read head of "ProzKraft.dat" file (number of experiment, ...)
    D2 = open(Datei, "r")
    Zeile = D2.readline()
    D2.close()

    # Find number of experiment
    VStart = Zeile.find(':') + 1
    VEnde = Zeile.find(' ')
    Versuchsnummer = Zeile[VStart:VEnde]
    CStart = Zeile.find(':', Zeile.find(':') + 1) + 1
    CEnd = Zeile.find('\n')
    Counter = Zeile[CStart:CEnd]
    Experiments[j] = str(Versuchsnummer + '_' + Counter)

    # Find row number of experiment
    for cell in MainSheet['A']:
        # if cell.value == int(Versuchsnummer):
        if cell.value == str(Versuchsnummer + '_' + Counter):
            ZeileExcel = cell.row

    # Find cutting speed
    vc = float(MainSheet.cell(row=ZeileExcel, column=15).value)

    # Find feed
    f = float(MainSheet.cell(row=ZeileExcel, column=12).value)

    # Find Wandstaerke
    Wandstaerke = float(MainSheet.cell(row=ZeileExcel, column=18).value) * 1000

    # Find # force measurements
    Repetitions = int(MainSheet.cell(row=ZeileExcel, column=16).value)

    #Repetitions = 1

    # Initalise array size depending on # Repetions
    AvgFc = np.zeros((len(Auswertegrenzen), Repetitions))
    AvgFf = np.zeros((len(Auswertegrenzen), Repetitions))
    AvgFp = np.zeros((len(Auswertegrenzen), Repetitions))
    SigFc = np.zeros((len(Auswertegrenzen), Repetitions))
    SigFf = np.zeros((len(Auswertegrenzen), Repetitions))
    SigFp = np.zeros((len(Auswertegrenzen), Repetitions))
    # Time points of force evaluation
    TimeForce = np.zeros((len(Auswertegrenzen), (4 * Repetitions)))

    #
    Daten = np.loadtxt(Datei, skiprows=2)

    AnzWerte = Daten[:, 1].size

    Data = []

    for file in os.listdir(Path_eval_files):
        if ('.txt' in file) and (('ProcessForceEvaluation_' + str(Experiments[j])) in file):
            Data.append(os.path.join(Path_eval_files, file))

    Data_row = np.genfromtxt(Data[0], dtype='str')

    if not EvalDetails.cell(row=ZeileExcel, column=2).value == None and DriftCompensation != 0:
        Time, Fc, Ff, Fp = InitaliseData(Daten)
        m_drift, q_drift, TimeDrift = Load_compensation_txt(
            Data_row, Repetitions)

        Fc, Ff, Fp, Time = Cut(Fc, Ff, Fp, TimeDrift, Time)
        Fc, Ff, Fp = Compensate(Fc, Ff, Fp, Time, m_drift, q_drift)

    elif EvalDetails.cell(row=ZeileExcel, column=2).value == None and DriftCompensation != 0:
        NoDrift = input(('No drift compensation available for' +
                         str(Versuchsnummer) + "_" + str(Counter) + ' experiment! Please confirm with [Y] '))
        if NoDrift == "Y" or NoDrift == "y":
            print("Thank you!")
        else:
            print(
                ('Experiment ' + str(Versuchsnummer) + '\_' + str(Counter) + ' is not evaluated!'))
            continue

    AvgFc, SigFc, AvgFf, SigFf, AvgFp, SigFp, TimeForce = Load_ForceEval_txt(
        Data_row, Repetitions)

    # Choose interactively borders for force evaluation
    fig, ax = plt.subplots(constrained_layout=True)
    # plt.xlim(TimeDrift[j, 2], TimeDrift[j, 4])
    ax.plot(Time, Fc, color="blue", linewidth=1.0, linestyle="-",
            label=r'Cutting force [N/mm]', zorder=2)
    ax.plot(Time, Ff, color="red", linewidth=1.0, linestyle="-",
            label=r'Feed force [N/mm]', zorder=2)
    ax.plot(Time, Fp, color="green",
            linewidth=1.0, linestyle="-", label=r'Passive force [N/mm]', zorder=2)

    # Plot evaluation points
    for i in range(Repetitions):
        Start = int(TimeForce[0, (4 * i + 1)])
        End = int(TimeForce[0, (4 * i + 3)])

        ax.axvline(x=Time[Start], color='k')
        ax.axvline(x=Time[End], color='k')
        ax.hlines(AvgFc[0, i], Time[Start], Time[End],
                  colors='k', zorder=1)
        ax.hlines(AvgFf[0, i], Time[Start],
                  Time[End], colors='k', zorder=1)
        ax.hlines(AvgFp[0, i], Time[Start],
                  Time[End], colors='k', zorder=1)

    # Plot settings
    ax.set_xlabel(r'Time [$s$]')
    ax.set_ylabel(r'Force [N/mm]')
    ax.set_title(r'Process forces, Experiment ' + Versuchsnummer + "\_" + Counter +
                 ', $v_c=$ ' + str(vc) + ' [m/min]' + ', $f=$ ' + str(f) + ' [mm]')
    ax.legend(loc="lower center")

    plt.savefig(r"Figures\EPS\Reload_" +
                Versuchsnummer + "_" + Counter + ".eps", dpi=300)
    plt.savefig(r"Figures\PNG\Reload_" +
                Versuchsnummer + "_" + Counter + ".png", dpi=300)

    plt.show(block=True)
    plt.close()

print('Reload done!')
